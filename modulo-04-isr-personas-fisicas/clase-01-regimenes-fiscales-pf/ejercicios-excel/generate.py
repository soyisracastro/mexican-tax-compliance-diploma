"""
Generador de ejercicios Excel para la Clase 01 del Módulo 4.

Requisitos: openpyxl >= 3.1
Ejecutar desde esta carpeta:
    python3 generate.py

Produce tres archivos xlsx:
    01-retencion-isr-mensual.xlsx
    02-resico-vs-612.xlsx
    03-arrendador-tres-opciones.xlsx

Cada archivo lleva 4 hojas:
    Instrucciones · Datos · Ejercicio · Solución
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

# ---------------------------------------------------------------------------
# Estilos compartidos (paleta TodoConta)
# ---------------------------------------------------------------------------
PRIMARY = "0B5FFF"
PRIMARY_SOFT = "EFF4FF"
MUTED = "F3F2EE"
BORDER = "D4D1C9"

FONT_TITLE = Font(name="Inter", size=16, bold=True, color="0A1628")
FONT_H2 = Font(name="Inter", size=13, bold=True, color=PRIMARY)
FONT_H3 = Font(name="Inter", size=11, bold=True, color="0A1628")
FONT_BODY = Font(name="Inter", size=11, color="0A1628")
FONT_BODY_MUTED = Font(name="Inter", size=10, color="555249")
FONT_HEADER = Font(name="Inter", size=10, bold=True, color="FFFFFF")
FONT_RESULT = Font(name="Inter", size=11, bold=True, color=PRIMARY)

FILL_PRIMARY = PatternFill("solid", fgColor=PRIMARY)
FILL_PRIMARY_SOFT = PatternFill("solid", fgColor=PRIMARY_SOFT)
FILL_MUTED = PatternFill("solid", fgColor=MUTED)
FILL_INPUT = PatternFill("solid", fgColor="FFF9C4")  # amarillo suave — celda a llenar

THIN = Side(border_style="thin", color=BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def set_column_widths(ws: Worksheet, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws: Worksheet, title: str, subtitle: str = ""):
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = FONT_BODY_MUTED
    ws.row_dimensions[1].height = 24


def write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list], *, money_cols: list[int] | None = None, percent_cols: list[int] | None = None):
    """Escribe tabla con header azul y bordes. Devuelve última fila escrita."""
    money_cols = money_cols or []
    percent_cols = percent_cols or []
    # Header
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    # Data
    for row_offset, row_values in enumerate(rows, 1):
        r = start_row + row_offset
        for col_idx, v in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.font = FONT_BODY
            cell.border = BORDER_ALL
            if col_idx in money_cols:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = ALIGN_RIGHT
            elif col_idx in percent_cols:
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
    return start_row + len(rows)


# ---------------------------------------------------------------------------
# Datos de referencia — Tarifas 2026
# ---------------------------------------------------------------------------
TARIFA_MENSUAL_2026 = [
    # (Límite inferior, Límite superior, Cuota fija, % excedente)
    (0.01, 844.59, 0.00, 0.0192),
    (844.60, 7165.92, 16.22, 0.0640),
    (7165.93, 12593.36, 420.79, 0.1088),
    (12593.37, 14632.96, 1011.05, 0.1600),
    (14632.97, 17523.45, 1337.38, 0.1792),
    (17523.46, 35299.65, 1852.05, 0.2136),
    (35299.66, 55640.58, 5651.62, 0.2352),
    (55640.59, 167424.91, 10434.61, 0.3000),
    (167424.92, 223233.22, 43969.91, 0.3200),
    (223233.23, 669699.66, 61828.55, 0.3400),
    (669699.67, 999999999.99, 213619.92, 0.3500),
]

TARIFA_ANUAL_2026 = [
    (0.01, 10134.92, 0.00, 0.0192),
    (10134.93, 85991.07, 194.59, 0.0640),
    (85991.08, 151120.33, 5049.50, 0.1088),
    (151120.34, 175595.60, 12132.58, 0.1600),
    (175595.61, 210281.40, 16008.60, 0.1792),
    (210281.41, 423595.89, 22224.52, 0.2136),
    (423595.90, 667687.06, 67811.43, 0.2352),
    (667687.07, 2009098.96, 125215.26, 0.3000),
    (2009098.97, 2678798.66, 527638.82, 0.3200),
    (2678798.67, 8036395.93, 741942.82, 0.3400),
    (8036395.94, 999999999.99, 2563435.86, 0.3500),
]

TASAS_RESICO_ANUAL = [
    # (Límite inferior, Límite superior, Tasa)
    (0.01, 300000.00, 0.0100),
    (300000.01, 600000.00, 0.0110),
    (600000.01, 1000000.00, 0.0150),
    (1000000.01, 2500000.00, 0.0200),
    (2500000.01, 3500000.00, 0.0250),
]

TASAS_RESICO_MENSUAL = [
    (0.01, 25000.00, 0.0100),
    (25000.01, 50000.00, 0.0110),
    (50000.01, 83333.33, 0.0150),
    (83333.34, 208333.33, 0.0200),
    (208333.34, 291666.67, 0.0250),
]


# ---------------------------------------------------------------------------
# Ejercicio 1 — Retención ISR mensual de sueldos
# ---------------------------------------------------------------------------
def build_ejercicio_1(path: Path):
    wb = Workbook()

    # ========= Hoja 1: Instrucciones =========
    ws = wb.active
    ws.title = "Instrucciones"
    set_column_widths(ws, [90])
    title_block(
        ws,
        "Ejercicio 1 — Retención ISR mensual de sueldos",
        "Art. 96 LISR · Tarifa mensual 2026 (Anexo 8 RMF)",
    )

    lines = [
        "",
        "OBJETIVO",
        "Construir una calculadora de retención mensual de ISR en Excel aplicando la tarifa progresiva del Art. 96 LISR.",
        "",
        "PASOS",
        "1. Ve a la hoja 'Datos' y revisa la tarifa mensual 2026. Cada renglón tiene: límite inferior, límite superior, cuota fija y % sobre excedente.",
        "2. Ve a la hoja 'Ejercicio'. Verás 5 casos con diferentes sueldos en la columna B (celdas amarillas = input).",
        "3. Para cada caso, llena las columnas C a I aplicando la fórmula:",
        "      ISR  =  (Sueldo − Límite inferior) × Tasa  +  Cuota fija",
        "4. Usa BUSCARV con aproximación (último argumento VERDADERO) sobre el rango de la tarifa para encontrar automáticamente el LI, la cuota fija y la tasa.",
        "   Tip: la tabla de BUSCARV debe tener el límite inferior en la primera columna.",
        "5. Calcula la carga efectiva (columna I) como ISR ÷ Sueldo.",
        "6. Compara tus resultados con la hoja 'Solución' al final.",
        "",
        "PISTAS DE FÓRMULAS",
        "  Límite inferior:  =BUSCARV(B4, Datos!$B$5:$E$15, 1, VERDADERO)",
        "  Cuota fija:       =BUSCARV(B4, Datos!$B$5:$E$15, 3, VERDADERO)",
        "  Tasa:             =BUSCARV(B4, Datos!$B$5:$E$15, 4, VERDADERO)",
        "  Excedente:        =B4 − C4",
        "  Impuesto marginal:=D4 × E4",
        "  ISR:              =F4 + G4",
        "  Carga efectiva:   =H4 / B4",
        "",
        "TIEMPO SUGERIDO: 15 minutos",
        "",
        "DISCUSIÓN EN PLENARIA",
        "¿Por qué un trabajador con $25,000 mensuales paga efectivo ~14% y uno con $200,000 paga ~27%? Identifica el principio constitucional en acción.",
    ]
    for i, txt in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=txt).alignment = ALIGN_LEFT
        if txt.isupper() and len(txt) > 2:
            ws.cell(row=i, column=1).font = FONT_H3

    # ========= Hoja 2: Datos =========
    ws = wb.create_sheet("Datos")
    set_column_widths(ws, [5, 18, 18, 16, 14])
    title_block(ws, "Tarifa mensual ISR 2026", "Anexo 8 RMF 2026 · DOF 28-dic-2025")

    headers = ["#", "Límite inferior", "Límite superior", "Cuota fija", "% excedente"]
    rows = [[i + 1, *row] for i, row in enumerate(TARIFA_MENSUAL_2026)]
    last = write_table(ws, 4, headers, rows, money_cols=[2, 3, 4], percent_cols=[5])

    ws.cell(row=last + 2, column=1, value="Factor de actualización 2026: 1.13213").font = FONT_BODY_MUTED

    # ========= Hoja 3: Ejercicio =========
    ws = wb.create_sheet("Ejercicio")
    set_column_widths(ws, [22, 14, 15, 13, 10, 15, 13, 14, 12])
    title_block(ws, "Cálculo de retención mensual — 5 casos", "Llena las celdas blancas usando las fórmulas indicadas en Instrucciones.")

    headers = [
        "Caso", "Sueldo (input)", "LI aplicable", "Excedente",
        "Tasa", "Imp. marginal", "Cuota fija", "ISR", "Carga efectiva",
    ]
    casos = [
        ("A — Operativo", 8000),
        ("B — Analista", 15000),
        ("C — Coordinador", 25000),
        ("D — Gerente", 50000),
        ("E — Directivo", 200000),
    ]

    # Header
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    for i, (nombre, sueldo) in enumerate(casos):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT

        # Input (amarillo)
        c = ws.cell(row=r, column=2, value=sueldo)
        c.fill = FILL_INPUT
        c.number_format = '"$"#,##0.00'
        c.alignment = ALIGN_RIGHT
        c.border = BORDER_ALL
        c.font = FONT_BODY

        # Resto en blanco — para que el alumno llene
        for col in range(3, 10):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER_ALL
            cell.alignment = ALIGN_RIGHT
            if col in (3, 6, 7, 8):
                cell.number_format = '"$"#,##0.00'
            elif col in (5, 9):
                cell.number_format = "0.00%"

    ws.cell(row=12, column=1, value="Celdas amarillas = input. Celdas blancas = a calcular.").font = FONT_BODY_MUTED

    # ========= Hoja 4: Solución =========
    ws = wb.create_sheet("Solución")
    set_column_widths(ws, [22, 14, 15, 13, 10, 15, 13, 14, 12])
    title_block(ws, "Solución — referencia con fórmulas", "Fórmulas con BUSCARV sobre la hoja Datos.")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    for i, (nombre, sueldo) in enumerate(casos):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT

        ws.cell(row=r, column=2, value=sueldo).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=2).fill = FILL_PRIMARY_SOFT

        # Fórmulas
        ws.cell(row=r, column=3, value=f"=VLOOKUP(B{r},Datos!$B$5:$E$15,1,TRUE)").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=f"=VLOOKUP(B{r},Datos!$B$5:$E$15,4,TRUE)").number_format = "0.00%"
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7, value=f"=VLOOKUP(B{r},Datos!$B$5:$E$15,3,TRUE)").number_format = '"$"#,##0.00'
        cell_isr = ws.cell(row=r, column=8, value=f"=F{r}+G{r}")
        cell_isr.number_format = '"$"#,##0.00'
        cell_isr.font = FONT_RESULT
        ws.cell(row=r, column=9, value=f"=H{r}/B{r}").number_format = "0.00%"

        for col in range(1, 10):
            ws.cell(row=r, column=col).border = BORDER_ALL
            if col > 1:
                ws.cell(row=r, column=col).alignment = ALIGN_RIGHT

    wb.save(path)
    print(f"✓ {path.name}")


# ---------------------------------------------------------------------------
# Ejercicio 2 — RESICO vs Actividad Empresarial (anual)
# ---------------------------------------------------------------------------
def build_ejercicio_2(path: Path):
    wb = Workbook()

    # ========= Hoja 1: Instrucciones =========
    ws = wb.active
    ws.title = "Instrucciones"
    set_column_widths(ws, [90])
    title_block(
        ws,
        "Ejercicio 2 — ¿RESICO o Actividad Empresarial?",
        "Comparativo anual para 5 clientes con distintos perfiles de deducciones",
    )

    lines = [
        "",
        "OBJETIVO",
        "Decidir qué régimen le conviene a cada cliente calculando el ISR anual en los dos escenarios.",
        "",
        "PASOS",
        "1. Hoja 'Datos': revisa la tarifa anual del Art. 152 LISR y las tasas anuales de RESICO (Art. 113-F).",
        "2. Hoja 'Ejercicio': tienes 5 clientes con sus ingresos y gastos anuales (celdas amarillas = input).",
        "3. Para cada cliente calcula:",
        "      Columna D — Base 612 = Ingresos − Gastos",
        "      Columna E — ISR 612 aplicando la tarifa anual (BUSCARV sobre Datos)",
        "      Columna F — ISR RESICO = Ingresos × Tasa RESICO (BUSCARV)",
        "      Columna G — Ahorro = ISR 612 − ISR RESICO (positivo = gana RESICO)",
        "      Columna H — Recomendación con SI()",
        "4. Observa en qué casos RESICO no conviene (muchos gastos) y cuándo es obligatorio (ingresos > $3.5 MDP).",
        "",
        "FÓRMULA CLAVE PARA ISR 612 ANUAL",
        "  LI:           =BUSCARV(D4, Datos!$B$5:$E$15, 1, VERDADERO)",
        "  Cuota fija:   =BUSCARV(D4, Datos!$B$5:$E$15, 3, VERDADERO)",
        "  Tasa:         =BUSCARV(D4, Datos!$B$5:$E$15, 4, VERDADERO)",
        "  ISR 612:      =(D4 − LI) × Tasa + Cuota fija",
        "",
        "FÓRMULA PARA ISR RESICO",
        "  ISR RESICO:   =B4 × BUSCARV(B4, Datos!$B$19:$D$23, 3, VERDADERO)",
        "",
        "FÓRMULA RECOMENDACIÓN (considerando límite RESICO de $3.5M)",
        '  =SI(B4>3500000, "Solo 612 (supera límite RESICO)", SI(G4>0, "RESICO", "612"))',
        "",
        "TIEMPO SUGERIDO: 20 minutos",
        "",
        "DISCUSIÓN EN PLENARIA",
        "¿A partir de qué % de gastos deja de convenir RESICO? ¿Cómo cambia la decisión si el cliente además es socio de una persona moral?",
    ]
    for i, txt in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=txt).alignment = ALIGN_LEFT
        if txt.isupper() and len(txt) > 2:
            ws.cell(row=i, column=1).font = FONT_H3

    # ========= Hoja 2: Datos =========
    ws = wb.create_sheet("Datos")
    set_column_widths(ws, [5, 20, 20, 16, 14])
    title_block(ws, "Tarifa anual Art. 152 + Tasas RESICO anual")

    ws["A3"] = "Tarifa anual 2026 (Art. 152 LISR)"
    ws["A3"].font = FONT_H3
    headers_anual = ["#", "Límite inferior", "Límite superior", "Cuota fija", "% excedente"]
    rows_anual = [[i + 1, *row] for i, row in enumerate(TARIFA_ANUAL_2026)]
    write_table(ws, 4, headers_anual, rows_anual, money_cols=[2, 3, 4], percent_cols=[5])

    ws["A17"] = "Tasas RESICO anual (Art. 113-F LISR)"
    ws["A17"].font = FONT_H3
    headers_resico = ["#", "Límite inferior", "Límite superior", "Tasa"]
    rows_resico = [[i + 1, *row] for i, row in enumerate(TASAS_RESICO_ANUAL)]
    write_table(ws, 18, headers_resico, rows_resico, money_cols=[2, 3], percent_cols=[4])

    # ========= Hoja 3: Ejercicio =========
    ws = wb.create_sheet("Ejercicio")
    set_column_widths(ws, [24, 16, 16, 16, 16, 16, 16, 30])
    title_block(ws, "5 clientes — ¿RESICO o 612?")

    headers = [
        "Cliente", "Ingresos anuales (input)", "Gastos anuales (input)",
        "Base 612", "ISR 612", "ISR RESICO", "Ahorro (612 − RESICO)", "Recomendación",
    ]

    clientes = [
        ("Arquitecto freelance (home office)", 1_200_000, 180_000),
        ("Dueño de taller mecánico", 1_200_000, 720_000),
        ("Consultora independiente", 800_000, 100_000),
        ("Comerciante mayorista", 4_200_000, 2_800_000),  # >$3.5M, solo 612
        ("Profesionista joven", 350_000, 50_000),
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    for i, (nombre, ing, gas) in enumerate(clientes):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL

        for col, val in [(2, ing), (3, gas)]:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = FILL_INPUT
            c.number_format = '"$"#,##0'
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_ALL
            c.font = FONT_BODY

        for col in range(4, 9):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER_ALL
            cell.alignment = ALIGN_RIGHT
            if col in (4, 5, 6, 7):
                cell.number_format = '"$"#,##0'
            elif col == 8:
                cell.alignment = ALIGN_LEFT

    ws.cell(row=12, column=1, value="Celdas amarillas = input. Celdas blancas = a calcular.").font = FONT_BODY_MUTED

    # ========= Hoja 4: Solución =========
    ws = wb.create_sheet("Solución")
    set_column_widths(ws, [24, 16, 16, 16, 16, 16, 16, 30])
    title_block(ws, "Solución — fórmulas completas")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    for i, (nombre, ing, gas) in enumerate(clientes):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL

        ws.cell(row=r, column=2, value=ing).number_format = '"$"#,##0'
        ws.cell(row=r, column=2).fill = FILL_PRIMARY_SOFT
        ws.cell(row=r, column=3, value=gas).number_format = '"$"#,##0'
        ws.cell(row=r, column=3).fill = FILL_PRIMARY_SOFT

        # Base 612 = ingresos − gastos
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = '"$"#,##0'
        # ISR 612 (tarifa anual)
        isr_612 = f"=(D{r}-VLOOKUP(D{r},Datos!$B$5:$E$15,1,TRUE))*VLOOKUP(D{r},Datos!$B$5:$E$15,4,TRUE)+VLOOKUP(D{r},Datos!$B$5:$E$15,3,TRUE)"
        ws.cell(row=r, column=5, value=isr_612).number_format = '"$"#,##0'
        # ISR RESICO (solo si <= 3.5M, sino N/A)
        isr_resico = f'=IF(B{r}>3500000,"N/A",B{r}*VLOOKUP(B{r},Datos!$B$19:$D$23,3,TRUE))'
        ws.cell(row=r, column=6, value=isr_resico).number_format = '"$"#,##0'
        # Ahorro
        ahorro = f'=IF(ISNUMBER(F{r}),E{r}-F{r},"—")'
        ws.cell(row=r, column=7, value=ahorro).number_format = '"$"#,##0'
        # Recomendación
        reco = f'=IF(B{r}>3500000,"Solo 612 (supera límite RESICO)",IF(G{r}>0,"RESICO",IF(G{r}<0,"612","Empate")))'
        cell_reco = ws.cell(row=r, column=8, value=reco)
        cell_reco.alignment = ALIGN_LEFT
        cell_reco.font = FONT_RESULT

        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER_ALL
            if col > 1 and col < 8:
                ws.cell(row=r, column=col).alignment = ALIGN_RIGHT

    wb.save(path)
    print(f"✓ {path.name}")


# ---------------------------------------------------------------------------
# Ejercicio 3 — Arrendador: 3 opciones
# ---------------------------------------------------------------------------
def build_ejercicio_3(path: Path):
    wb = Workbook()

    # ========= Hoja 1: Instrucciones =========
    ws = wb.active
    ws.title = "Instrucciones"
    set_column_widths(ws, [90])
    title_block(
        ws,
        "Ejercicio 3 — Arrendador: comprobadas vs ciega vs RESICO",
        "Elegir la opción más conveniente para cada arrendador",
    )

    lines = [
        "",
        "OBJETIVO",
        "Para cada arrendador, calcular el ISR anual en las tres opciones disponibles y recomendar la mejor.",
        "",
        "LAS TRES OPCIONES",
        "  A — Régimen 606 con DEDUCCIONES COMPROBADAS (Art. 115, párrafo 1)",
        "      Deduce gastos reales: predial, mantenimiento, intereses reales, depreciación 5%, seguros.",
        "  B — Régimen 606 con DEDUCCIÓN CIEGA (Art. 115, párrafo 2)",
        "      Deduce 35% de ingresos + predial pagado, sin otros comprobantes.",
        "  C — Régimen 626 RESICO (si ingresos ≤ $3.5 MDP)",
        "      Tasa anual de 1% a 2.5% sobre ingresos cobrados, sin deducciones.",
        "",
        "PASOS",
        "1. Hoja 'Datos': tarifa anual Art. 152 + tasas RESICO anual.",
        "2. Hoja 'Ejercicio': 3 arrendadores con sus datos (celdas amarillas = input).",
        "3. Para cada arrendador llena:",
        "      Base A (comprobadas) = Ingresos − Gastos reales − Depreciación",
        "      Base B (ciega)        = Ingresos − (Ingresos × 35%) − Predial",
        "      ISR A, ISR B aplicando tarifa anual (BUSCARV)",
        "      ISR C (RESICO)        = Ingresos × Tasa RESICO  (o N/A si >$3.5M)",
        "      Recomendación         = la opción de menor ISR",
        "",
        "TIEMPO SUGERIDO: 20 minutos",
        "",
        "DISCUSIÓN EN PLENARIA",
        "¿En qué escenario gana comprobadas sobre ciega? ¿A partir de qué nivel de ingresos RESICO deja de ser la obvia?",
    ]
    for i, txt in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=txt).alignment = ALIGN_LEFT
        if txt.isupper() and len(txt) > 2:
            ws.cell(row=i, column=1).font = FONT_H3

    # ========= Hoja 2: Datos =========
    ws = wb.create_sheet("Datos")
    set_column_widths(ws, [5, 20, 20, 16, 14])
    title_block(ws, "Tarifa anual Art. 152 + Tasas RESICO anual")

    ws["A3"] = "Tarifa anual 2026 (Art. 152 LISR)"
    ws["A3"].font = FONT_H3
    headers_anual = ["#", "Límite inferior", "Límite superior", "Cuota fija", "% excedente"]
    rows_anual = [[i + 1, *row] for i, row in enumerate(TARIFA_ANUAL_2026)]
    write_table(ws, 4, headers_anual, rows_anual, money_cols=[2, 3, 4], percent_cols=[5])

    ws["A17"] = "Tasas RESICO anual (Art. 113-F LISR)"
    ws["A17"].font = FONT_H3
    headers_resico = ["#", "Límite inferior", "Límite superior", "Tasa"]
    rows_resico = [[i + 1, *row] for i, row in enumerate(TASAS_RESICO_ANUAL)]
    write_table(ws, 18, headers_resico, rows_resico, money_cols=[2, 3], percent_cols=[4])

    # ========= Hoja 3: Ejercicio =========
    ws = wb.create_sheet("Ejercicio")
    widths = [22, 14, 14, 14, 14, 12, 12, 12, 12, 18]
    set_column_widths(ws, widths)
    title_block(ws, "3 arrendadores — 3 opciones cada uno")

    headers = [
        "Arrendador",
        "Ingresos anuales",
        "Gastos reales",
        "Depreciación 5%",
        "Predial anual",
        "ISR A (comprob.)",
        "ISR B (ciega 35%)",
        "ISR C (RESICO)",
        "Mejor opción",
        "ISR mínimo",
    ]

    arrendadores = [
        # (nombre, ingresos, gastos reales, depreciación, predial)
        ("Doña Carmen — 2 casas", 444_000, 20_000, 25_000, 9_700),
        ("Don Miguel — 1 local", 300_000, 5_000, 12_000, 4_500),
        ("Sra. Ana — 4 deptos con hipoteca", 600_000, 95_000, 40_000, 15_000),
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    for i, (nombre, ing, gas, dep, pre) in enumerate(arrendadores):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL

        for col, val in [(2, ing), (3, gas), (4, dep), (5, pre)]:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = FILL_INPUT
            c.number_format = '"$"#,##0'
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_ALL
            c.font = FONT_BODY

        for col in range(6, 11):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER_ALL
            cell.alignment = ALIGN_RIGHT
            if col != 9:
                cell.number_format = '"$"#,##0'
            else:
                cell.alignment = ALIGN_LEFT

    ws.cell(row=10, column=1, value="Celdas amarillas = input. Celdas blancas = a calcular.").font = FONT_BODY_MUTED

    # ========= Hoja 4: Solución =========
    ws = wb.create_sheet("Solución")
    set_column_widths(ws, widths)
    title_block(ws, "Solución — fórmulas completas")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    for i, (nombre, ing, gas, dep, pre) in enumerate(arrendadores):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL

        # Inputs
        for col, val in [(2, ing), (3, gas), (4, dep), (5, pre)]:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = FILL_PRIMARY_SOFT
            c.number_format = '"$"#,##0'
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_ALL

        # Base A (comprobadas) = Ingresos − Gastos − Depreciación − Predial
        base_a = f"(B{r}-C{r}-D{r}-E{r})"
        # Base B (ciega) = Ingresos × 65% − Predial
        base_b = f"(B{r}*0.65-E{r})"

        # ISR A
        isr_a = (
            f"=MAX(0,({base_a}-VLOOKUP({base_a},Datos!$B$5:$E$15,1,TRUE))"
            f"*VLOOKUP({base_a},Datos!$B$5:$E$15,4,TRUE)"
            f"+VLOOKUP({base_a},Datos!$B$5:$E$15,3,TRUE))"
        )
        # ISR B
        isr_b = (
            f"=MAX(0,({base_b}-VLOOKUP({base_b},Datos!$B$5:$E$15,1,TRUE))"
            f"*VLOOKUP({base_b},Datos!$B$5:$E$15,4,TRUE)"
            f"+VLOOKUP({base_b},Datos!$B$5:$E$15,3,TRUE))"
        )
        # ISR C (RESICO si <= 3.5M)
        isr_c = f'=IF(B{r}>3500000,"N/A",B{r}*VLOOKUP(B{r},Datos!$B$19:$D$23,3,TRUE))'

        ws.cell(row=r, column=6, value=isr_a).number_format = '"$"#,##0'
        ws.cell(row=r, column=7, value=isr_b).number_format = '"$"#,##0'
        ws.cell(row=r, column=8, value=isr_c).number_format = '"$"#,##0'

        # Mejor opción
        mejor = (
            f'=IF(ISNUMBER(H{r}),'
            f'IF(MIN(F{r},G{r},H{r})=H{r},"C — RESICO",IF(MIN(F{r},G{r})=F{r},"A — Comprobadas","B — Ciega 35%")),'
            f'IF(F{r}<G{r},"A — Comprobadas","B — Ciega 35%"))'
        )
        cell_mejor = ws.cell(row=r, column=9, value=mejor)
        cell_mejor.alignment = ALIGN_LEFT
        cell_mejor.font = FONT_RESULT

        # ISR mínimo
        isr_min = f'=IF(ISNUMBER(H{r}),MIN(F{r},G{r},H{r}),MIN(F{r},G{r}))'
        ws.cell(row=r, column=10, value=isr_min).number_format = '"$"#,##0'

        for col in range(1, 11):
            ws.cell(row=r, column=col).border = BORDER_ALL
            if col > 1 and col != 9:
                ws.cell(row=r, column=col).alignment = ALIGN_RIGHT

    wb.save(path)
    print(f"✓ {path.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    here = Path(__file__).parent
    build_ejercicio_1(here / "01-retencion-isr-mensual.xlsx")
    build_ejercicio_2(here / "02-resico-vs-612.xlsx")
    build_ejercicio_3(here / "03-arrendador-tres-opciones.xlsx")
    print("\nListo. 3 ejercicios generados.")


if __name__ == "__main__":
    main()
