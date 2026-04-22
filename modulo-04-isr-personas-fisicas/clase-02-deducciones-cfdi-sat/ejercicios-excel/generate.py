"""
Generador de ejercicios Excel para la Clase 02 del Módulo 4.

Requisitos: openpyxl >= 3.1
Ejecutar desde esta carpeta:
    python3 generate.py

Produce dos archivos xlsx:
    01-deducciones-personales-tope.xlsx
    02-declaracion-anual-pf.xlsx

Cada archivo lleva 4 hojas:
    Instrucciones · Datos · Ejercicio · Solución

Tarifa 2026: Anexo 8 RMF 2026 (DOF 28-dic-2025), factor 1.13213.
Nota: NO usar la tarifa del archivo de clase-01 (tiene factor incorrecto).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

# ---------------------------------------------------------------------------
# Estilos compartidos (paleta TodoConta)
# ---------------------------------------------------------------------------
PRIMARY = "0B5FFF"
PRIMARY_SOFT = "EFF4FF"
MUTED = "F3F2EE"
BORDER = "D4D1C9"
YELLOW_INPUT = "FFF9C4"
GREEN_SOFT = "E8F5E9"
RED_SOFT = "FFEBEE"

FONT_TITLE = Font(name="Inter", size=16, bold=True, color="0A1628")
FONT_H2 = Font(name="Inter", size=13, bold=True, color=PRIMARY)
FONT_H3 = Font(name="Inter", size=11, bold=True, color="0A1628")
FONT_BODY = Font(name="Inter", size=11, color="0A1628")
FONT_BODY_MUTED = Font(name="Inter", size=10, color="555249")
FONT_HEADER = Font(name="Inter", size=10, bold=True, color="FFFFFF")
FONT_RESULT = Font(name="Inter", size=11, bold=True, color=PRIMARY)
FONT_NOTE = Font(name="Inter", size=10, italic=True, color="B00020")

FILL_PRIMARY = PatternFill("solid", fgColor=PRIMARY)
FILL_PRIMARY_SOFT = PatternFill("solid", fgColor=PRIMARY_SOFT)
FILL_MUTED = PatternFill("solid", fgColor=MUTED)
FILL_INPUT = PatternFill("solid", fgColor=YELLOW_INPUT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN_SOFT)
FILL_RED = PatternFill("solid", fgColor=RED_SOFT)

THIN = Side(border_style="thin", color=BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def set_column_widths(ws: Worksheet, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws: Worksheet, title: str, subtitle: str = ""):
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = FONT_BODY_MUTED
    ws.row_dimensions[1].height = 26


def write_table(
    ws: Worksheet,
    start_row: int,
    headers: list[str],
    rows: list[list],
    *,
    money_cols: list[int] | None = None,
    percent_cols: list[int] | None = None,
):
    money_cols = money_cols or []
    percent_cols = percent_cols or []
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    for row_offset, row_values in enumerate(rows, 1):
        r = start_row + row_offset
        for col_idx, v in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.font = FONT_BODY
            cell.border = BORDER_ALL
            if col_idx in money_cols:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = ALIGN_RIGHT
            elif col_idx in percent_cols:
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
    return start_row + len(rows)


def input_cell(ws, row, col, value, *, money=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = FILL_INPUT
    c.number_format = '"$"#,##0' if money else "@"
    c.alignment = ALIGN_RIGHT
    c.border = BORDER_ALL
    c.font = FONT_BODY
    return c


def blank_cell(ws, row, col, *, money=True, pct=False):
    c = ws.cell(row=row, column=col)
    c.border = BORDER_ALL
    c.alignment = ALIGN_RIGHT
    if pct:
        c.number_format = "0.00%"
    elif money:
        c.number_format = '"$"#,##0.00'
    return c


# ---------------------------------------------------------------------------
# Datos de referencia — Tarifas 2026 CORREGIDAS (Anexo 8 RMF 2026, factor 1.13213)
# ---------------------------------------------------------------------------
TARIFA_ANUAL_2026 = [
    # (Límite inferior, Límite superior, Cuota fija, % excedente)
    (0.01, 10_134.92, 0.00, 0.0192),
    (10_134.93, 85_991.07, 194.59, 0.0640),
    (85_991.08, 151_120.33, 5_049.50, 0.1088),
    (151_120.34, 175_595.60, 12_132.58, 0.1600),
    (175_595.61, 210_281.40, 16_008.60, 0.1792),
    (210_281.41, 423_595.89, 22_224.52, 0.2136),
    (423_595.90, 667_687.06, 67_811.43, 0.2352),
    (667_687.07, 1_276_925.98, 125_215.26, 0.3000),
    (1_276_925.99, 1_702_567.97, 307_989.62, 0.3200),
    (1_702_567.98, 5_107_703.90, 448_131.88, 0.3400),
    (5_107_703.91, 999_999_999.99, 1_606_771.10, 0.3500),
]

# Constantes 2026
UMA_DIARIA_ENE = 113.14
UMA_ANUAL = 42_794.64          # publicado INEGI; base para tope global
TOPE_UMA_5 = 5 * UMA_ANUAL     # $213,973.20

SMG_DIARIO = 315.04
SMG_ANUAL = SMG_DIARIO * 365   # $114,989.60
TOPE_SMG_5 = 5 * SMG_ANUAL     # $574,948.00  — Art. 151 Frac. V


# ---------------------------------------------------------------------------
# Ejercicio 1 — Deducciones personales y tope global
# ---------------------------------------------------------------------------
def build_ejercicio_1(path: Path):
    wb = Workbook()

    # ===== Instrucciones =====
    ws = wb.active
    ws.title = "Instrucciones"
    set_column_widths(ws, [95])
    title_block(
        ws,
        "Ejercicio 1 — Deducciones Personales y Tope Global 2026",
        "Arts. 151 y 185 LISR · Decreto colegiaturas DOF 26-dic-2013",
    )

    lines = [
        "",
        "OBJETIVO",
        "Clasificar gastos, aplicar los límites específicos de cada fracción (especialmente el límite SMG del Art. 151 Frac. V) "
        "y calcular el tope global de deducciones personales para cuatro perfiles distintos.",
        "",
        "REGLAS CLAVE",
        "  • Tope global = el MENOR de: (a) 5 UMAs anuales = $213,973.20 ó (b) 15% de los ingresos acumulables del ejercicio.",
        "  • Colegiaturas (Decreto 2013) quedan FUERA del tope global — se suman al final.",
        "  • Art. 151 Frac. V — PPR/PPAP: límite = el MENOR de (a) 10% de ingresos ó (b) 5 SMG anuales ($574,948.00).",
        "    OJO: El límite es 5 SALARIOS MÍNIMOS, NO 5 UMAs. Error frecuente en muchos materiales de capacitación.",
        "  • Gastos con clave CFDI incorrecta (ej. G03 en lugar de D01) NO son deducibles como personales.",
        "",
        "PASOS",
        "1. Ve a la hoja 'Datos' — revisa los catálogos de referencia (UMA, SMG, colegiaturas).",
        "2. Ve a la hoja 'Ejercicio'. Tienes 4 perfiles con sus gastos (celdas amarillas = input ya proporcionado).",
        "3. Para cada perfil calcula en las celdas blancas:",
        "      Col H — PPR con límite: =MIN(G?, B?*10%, 574948)",
        "             donde G? = aportación bruta declarada, B? = ingresos anuales",
        "      Col I — Subtotal Arts. 151+185: suma de D?+E?+F?+H? (sin colegiaturas)",
        "      Col J — Tope 15%: =B?*15%",
        "      Col K — Tope 5 UMAs: =$213,973.20 (constante en hoja Datos)",
        "      Col L — Tope aplicable: =MIN(J?,K?)",
        "      Col M — Deducción Arts. 151+185: =MIN(I?,L?)  ← la que entra a la declaración",
        "      Col N — Deducción TOTAL: =M?+C?  ← suma colegiaturas (fuera de tope)",
        "4. Compara con la hoja 'Solución' al final.",
        "",
        "TIEMPO SUGERIDO: 20 minutos",
        "",
        "DISCUSIÓN EN PLENARIA",
        "¿A partir de qué nivel de ingresos el tope global deja de ser el 15% y pasa a ser fijo en $213,973?",
        "¿Por qué alguien con $3,000,000 de ingreso puede ser que sus deducciones personales no le signifiquen gran ahorro?",
    ]
    for i, txt in enumerate(lines, 3):
        c = ws.cell(row=i, column=1, value=txt)
        c.alignment = ALIGN_LEFT
        if txt.startswith("OBJETIVO") or txt.startswith("REGLAS") or txt.startswith("PASOS") or txt.startswith("TIEMPO") or txt.startswith("DISCUSIÓN"):
            c.font = FONT_H3
        elif txt.startswith("    OJO"):
            c.font = FONT_NOTE
        else:
            c.font = FONT_BODY

    # ===== Datos =====
    ws = wb.create_sheet("Datos")
    set_column_widths(ws, [5, 20, 20, 16, 14])
    title_block(ws, "Catálogos de referencia 2026", "Anexo 8 RMF 2026 · DOF 28-dic-2025")

    ws["A3"] = "Tarifa anual 2026 (Art. 152 LISR)"
    ws["A3"].font = FONT_H3
    headers_anual = ["#", "Límite inferior", "Límite superior", "Cuota fija", "% excedente"]
    rows_anual = [[i + 1, *row] for i, row in enumerate(TARIFA_ANUAL_2026)]
    last = write_table(ws, 4, headers_anual, rows_anual, money_cols=[2, 3, 4], percent_cols=[5])

    ws.cell(row=last + 2, column=1, value="Constantes 2026").font = FONT_H3
    constantes = [
        ("UMA diaria (enero 2026)", UMA_DIARIA_ENE),
        ("UMA anual 2026", UMA_ANUAL),
        ("5 UMAs anuales (tope global Arts. 151+185)", TOPE_UMA_5),
        ("SMG diario 2026", SMG_DIARIO),
        ("SMG anual 2026 (365 días)", SMG_ANUAL),
        ("5 SMG anuales — límite Art. 151 Frac. V (PPR/PPAP)", TOPE_SMG_5),
    ]
    for j, (etiqueta, valor) in enumerate(constantes):
        r = last + 3 + j
        ws.cell(row=r, column=1, value=etiqueta).font = FONT_BODY
        c = ws.cell(row=r, column=2, value=valor)
        c.number_format = '"$"#,##0.00'
        c.font = FONT_RESULT
        c.border = BORDER_ALL

    ws.cell(row=last + 10, column=1, value="Límites de colegiaturas (Decreto DOF 26-dic-2013)").font = FONT_H3
    colegiaturas = [
        ("Preescolar", 14_200),
        ("Primaria", 12_900),
        ("Secundaria", 19_900),
        ("Profesional técnico", 17_100),
        ("Bachillerato / preparatoria", 24_500),
        ("Licenciatura y superior", 0),
    ]
    for j, (nivel, limite) in enumerate(colegiaturas):
        r = last + 11 + j
        ws.cell(row=r, column=1, value=nivel).font = FONT_BODY
        c = ws.cell(row=r, column=2, value=limite if limite else "No aplica el decreto")
        if limite:
            c.number_format = '"$"#,##0'
        c.font = FONT_BODY
        c.border = BORDER_ALL

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    # ===== Ejercicio =====
    ws = wb.create_sheet("Ejercicio")
    set_column_widths(ws, [26, 14, 14, 13, 13, 13, 13, 15, 14, 12, 12, 13, 16, 16])

    title_block(
        ws,
        "4 perfiles — Deducciones personales 2026",
        "Celdas amarillas = input ya dado. Celdas blancas = a calcular.",
    )

    headers = [
        "Perfil",
        "Ingresos anuales",
        "Colegiaturas (D10)\n[fuera de tope]",
        "Médicos/hospital\n(D01)",
        "Seguros GMM\n(D07)",
        "Intereses\nhipoteca (D05)",
        "PPR bruto\ndeclarado (D06)",
        "PPR con límite\n(Min 10%/5SMG)",
        "Subtotal\nArts. 151+185",
        "Tope 15%\ningreso",
        "Tope 5 UMAs\n$213,973.20",
        "Tope que\naplica (Min)",
        "Deducción\nArts.151+185",
        "Deducción\nTOTAL",
    ]

    perfiles = [
        # (nombre, ingresos, colegiaturas, médicos, seguros, hipoteca, PPR_bruto)
        ("Sofía — Asalariada $350K", 350_000, 24_500, 45_000, 18_000, 0, 35_000),
        ("Roberto — Contador $937K", 937_000, 24_500, 60_000, 24_000, 28_000, 90_000),
        ("Gabriela — Arrendadora $180K", 180_000, 0, 8_000, 12_000, 0, 18_000),
        ("Miguel — RESICO $480K", 480_000, 19_900, 35_000, 22_000, 0, 48_000),
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 40

    for i, (nombre, ing, col, med, seg, hip, ppr) in enumerate(perfiles):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 18

        for col_idx, val in [(2, ing), (3, col), (4, med), (5, seg), (6, hip), (7, ppr)]:
            input_cell(ws, r, col_idx, val)

        # Celdas a calcular (blancas)
        for col_idx in range(8, 15):
            blank_cell(ws, r, col_idx)

    ws.cell(row=11, column=1,
            value="⚠ Art. 151 Frac. V: el límite PPR es 5 SMG anuales ($574,948), NO 5 UMAs. Ver hoja Datos.").font = FONT_NOTE
    ws.merge_cells("A11:G11")
    ws.cell(row=12, column=1,
            value="⚠ Colegiaturas (col C) se suman DESPUÉS del tope — van directamente a la deducción total.").font = FONT_NOTE
    ws.merge_cells("A12:G12")

    # ===== Solución =====
    ws = wb.create_sheet("Solución")
    set_column_widths(ws, [26, 14, 14, 13, 13, 13, 13, 15, 14, 12, 12, 13, 16, 16])
    title_block(ws, "Solución — fórmulas con las reglas correctas", "")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 40

    for i, (nombre, ing, col, med, seg, hip, ppr) in enumerate(perfiles):
        r = 5 + i
        ws.cell(row=r, column=1, value=nombre).font = FONT_BODY
        ws.cell(row=r, column=1).border = BORDER_ALL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 18

        # Inputs (azul suave)
        for col_idx, val in [(2, ing), (3, col), (4, med), (5, seg), (6, hip), (7, ppr)]:
            c = ws.cell(row=r, column=col_idx, value=val)
            c.fill = FILL_PRIMARY_SOFT
            c.number_format = '"$"#,##0'
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_ALL
            c.font = FONT_BODY

        # H — PPR con límite: MIN(G, B*10%, 574948)
        f_ppr = f"=MIN(G{r},B{r}*0.10,{TOPE_SMG_5})"
        ws.cell(row=r, column=8, value=f_ppr).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=8).border = BORDER_ALL
        ws.cell(row=r, column=8).alignment = ALIGN_RIGHT

        # I — Subtotal Arts. 151+185: D+E+F+H
        f_sub = f"=D{r}+E{r}+F{r}+H{r}"
        ws.cell(row=r, column=9, value=f_sub).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=9).border = BORDER_ALL
        ws.cell(row=r, column=9).alignment = ALIGN_RIGHT

        # J — Tope 15%
        f_15 = f"=B{r}*0.15"
        ws.cell(row=r, column=10, value=f_15).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=10).border = BORDER_ALL
        ws.cell(row=r, column=10).alignment = ALIGN_RIGHT

        # K — Tope 5 UMAs (constante)
        ws.cell(row=r, column=11, value=TOPE_UMA_5).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=11).border = BORDER_ALL
        ws.cell(row=r, column=11).alignment = ALIGN_RIGHT

        # L — Tope aplicable = MIN(J, K)
        f_tope = f"=MIN(J{r},K{r})"
        ws.cell(row=r, column=12, value=f_tope).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=12).border = BORDER_ALL
        ws.cell(row=r, column=12).alignment = ALIGN_RIGHT

        # M — Deducción Arts. 151+185 = MIN(I, L)
        f_ded = f"=MIN(I{r},L{r})"
        cell_ded = ws.cell(row=r, column=13, value=f_ded)
        cell_ded.number_format = '"$"#,##0.00'
        cell_ded.border = BORDER_ALL
        cell_ded.alignment = ALIGN_RIGHT
        cell_ded.font = FONT_RESULT

        # N — Deducción TOTAL = M + C (colegiaturas)
        f_total = f"=M{r}+C{r}"
        cell_total = ws.cell(row=r, column=14, value=f_total)
        cell_total.number_format = '"$"#,##0.00'
        cell_total.border = BORDER_ALL
        cell_total.alignment = ALIGN_RIGHT
        cell_total.font = FONT_RESULT
        cell_total.fill = FILL_GREEN

    ws.cell(row=11, column=1,
            value="Fórmula PPR (col H): =MIN(G,B*10%,574948) — límite es 5 SMG anuales, NOT 5 UMAs").font = FONT_NOTE
    ws.merge_cells("A11:N11")

    wb.save(path)
    print(f"✓ {path.name}")


# ---------------------------------------------------------------------------
# Ejercicio 2 — Declaración Anual de Personas Físicas 2026
# ---------------------------------------------------------------------------
def build_ejercicio_2(path: Path):
    wb = Workbook()

    # ===== Instrucciones =====
    ws = wb.active
    ws.title = "Instrucciones"
    set_column_widths(ws, [95])
    title_block(
        ws,
        "Ejercicio 2 — Declaración Anual ISR Personas Físicas 2026",
        "Art. 150 LISR · Tarifa anual Art. 152 (Anexo 8 RMF 2026)",
    )

    lines = [
        "",
        "OBJETIVO",
        "Construir la declaración anual de ISR para dos perfiles distintos: un asalariado con retenciones "
        "de patrón y un profesionista independiente con pagos provisionales propios. "
        "Aplicar la tarifa anual 2026 y determinar si hay ISR a cargo o saldo a favor.",
        "",
        "PERFIL A — JUAN (Asalariado, Capítulo I LISR)",
        "  · Un solo patrón durante el año — obligado a declarar por tener ingresos > $400,000 (Art. 150 LISR).",
        "  · El patrón retuvo ISR mensual durante el año (datos dados).",
        "  · Juan tiene deducciones personales: gastos médicos y primas de seguro de gastos médicos.",
        "",
        "PERFIL B — MARÍA (Contadora independiente, Capítulo II Sección I LISR)",
        "  · Ingresos de honorarios cobrados durante el año.",
        "  · Realizó pagos provisionales mensuales propios (Art. 106 LISR).",
        "  · Tiene gastos deducibles del negocio (indispensables) y deducciones personales.",
        "  · Aportó a PPR (Plan Personal de Retiro) — verificar límite Art. 151 Frac. V.",
        "",
        "MECÁNICA GENERAL DE LA DECLARACIÓN ANUAL",
        "  1. Ingresos acumulables del ejercicio",
        "  2. Menos: Deducciones propias del régimen (solo aplica para Capítulo II)",
        "  3. = Base gravable antes de deducciones personales",
        "  4. Menos: Deducciones personales (Arts. 151+185 + colegiaturas, sujeto a tope global)",
        "  5. = Base gravable final",
        "  6. Aplicar tarifa anual Art. 152 LISR 2026 (BUSCARV sobre hoja Datos)",
        "  7. Menos: retenciones por patrón (Juan) ó pagos provisionales propios (María)",
        "  8. = ISR a cargo (+) ó saldo a favor (−)",
        "",
        "FÓRMULA TARIFA ANUAL (BUSCARV)",
        "  LI aplicable:  =BUSCARV(BaseGravable, Datos!$B$5:$E$15, 1, VERDADERO)",
        "  Cuota fija:    =BUSCARV(BaseGravable, Datos!$B$5:$E$15, 3, VERDADERO)",
        "  Tasa:          =BUSCARV(BaseGravable, Datos!$B$5:$E$15, 4, VERDADERO)",
        "  ISR:           =(BaseGravable − LI) × Tasa + Cuota fija",
        "",
        "TIEMPO SUGERIDO: 25 minutos",
        "",
        "DISCUSIÓN EN PLENARIA",
        "¿Por qué María tiene más formas de reducir su base que Juan? ¿Qué herramienta del SAT les permite a ambos",
        "verificar antes de declarar que los CFDIs de sus deducciones ya están en el sistema?",
    ]
    for i, txt in enumerate(lines, 3):
        c = ws.cell(row=i, column=1, value=txt)
        c.alignment = ALIGN_LEFT
        if any(txt.startswith(k) for k in ["OBJETIVO", "PERFIL", "MECÁNICA", "FÓRMULA", "TIEMPO", "DISCUSIÓN"]):
            c.font = FONT_H3
        elif "OJO" in txt or "·" in txt[:4]:
            c.font = FONT_BODY
        else:
            c.font = FONT_BODY

    # ===== Datos =====
    ws = wb.create_sheet("Datos")
    set_column_widths(ws, [5, 20, 20, 16, 14])
    title_block(ws, "Tarifa anual Art. 152 LISR 2026", "Anexo 8 RMF 2026 · DOF 28-dic-2025 · Factor 1.13213")

    ws["A3"] = "Tarifa anual 2026"
    ws["A3"].font = FONT_H3
    headers_anual = ["#", "Límite inferior", "Límite superior", "Cuota fija", "% excedente"]
    rows_anual = [[i + 1, *row] for i, row in enumerate(TARIFA_ANUAL_2026)]
    last = write_table(ws, 4, headers_anual, rows_anual, money_cols=[2, 3, 4], percent_cols=[5])

    ws.cell(row=last + 2, column=1, value="Constantes para deducciones personales").font = FONT_H3
    constantes = [
        ("5 UMAs anuales — tope global Arts.151+185", TOPE_UMA_5),
        ("5 SMG anuales — límite Art.151 Frac. V (PPR)", TOPE_SMG_5),
    ]
    for j, (etiqueta, valor) in enumerate(constantes):
        r = last + 3 + j
        ws.cell(row=r, column=1, value=etiqueta).font = FONT_BODY
        c = ws.cell(row=r, column=2, value=valor)
        c.number_format = '"$"#,##0.00'
        c.font = FONT_RESULT
        c.border = BORDER_ALL

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    # ===== Ejercicio =====
    ws = wb.create_sheet("Ejercicio")
    set_column_widths(ws, [42, 20, 20])
    title_block(ws, "Declaración anual 2026 — Perfil A: Juan (Asalariado)", "")

    # --- JUAN ---
    ws["A3"] = "PERFIL A — JUAN PÉREZ (Capítulo I · Sueldos y Salarios)"
    ws["A3"].font = FONT_H2

    juan_rows = [
        ("Ingresos gravados por nómina (anual)", 650_000, True),
        ("Ingresos exentos (aguinaldo, prima vacacional)", None, False),
        ("Ingresos acumulables del ejercicio", None, False),
        ("", None, False),
        ("--- DEDUCCIONES PERSONALES ---", None, False),
        ("Gastos médicos y hospitalarios (D01)", 30_000, True),
        ("Primas seguro gastos médicos (D07)", 15_000, True),
        ("Aportaciones PPR — bruto declarado (D06)", 25_000, True),
        ("PPR con límite Art. 151 Frac. V [MIN(bruto, 10% ing, 5SMG)]", None, False),
        ("Subtotal deducciones Arts. 151+185", None, False),
        ("Tope 15% sobre ingresos acumulables", None, False),
        ("Tope global (MIN entre 15% y 5 UMAs = $213,973.20)", None, False),
        ("Deducciones personales que aplican", None, False),
        ("", None, False),
        ("--- CÁLCULO ISR ---", None, False),
        ("Base gravable (ingresos acum. − ded. personales)", None, False),
        ("LI aplicable (BUSCARV con VERDADERO)", None, False),
        ("Excedente sobre LI", None, False),
        ("Tasa marginal", None, False),
        ("Impuesto marginal (excedente × tasa)", None, False),
        ("Cuota fija (BUSCARV)", None, False),
        ("ISR determinado", None, False),
        ("Menos: retenciones por patrón (dato dado)", 122_000, True),
        ("ISR A CARGO (+) / SALDO A FAVOR (−)", None, False),
    ]

    for j, (etiqueta, valor, es_input) in enumerate(juan_rows):
        r = 4 + j
        ws.row_dimensions[r].height = 16
        c_label = ws.cell(row=r, column=1, value=etiqueta)
        c_label.font = FONT_BODY if etiqueta and not etiqueta.startswith("---") else FONT_H3
        c_label.alignment = ALIGN_LEFT
        c_label.border = BORDER_ALL if etiqueta and not etiqueta.startswith("---") else None

        if es_input and valor is not None:
            input_cell(ws, r, 2, valor)
        elif etiqueta and not etiqueta.startswith("---") and etiqueta != "":
            blank_cell(ws, r, 2)

        # Col C = nota
        notas = {
            4: "Art. 150 LISR — obligado a declarar (> $400,000)",
            5: "Art. 93 XIV — exenciones laborales (no acumulables)",
            6: "Suma fila 4 + 5 (pero las exentas ya no entran)",
            14: "Colegiaturas: N/A para Juan (no tiene hijos en edad escolar)",
            23: "Dato: el patrón retuvo esta cantidad durante el año (Art. 96)",
            24: "Si positivo: Juan paga en abril. Si negativo: solicita devolución.",
        }
        nota_row = r - 3
        if nota_row in notas:
            ws.cell(row=r, column=3, value=notas[nota_row]).font = FONT_BODY_MUTED

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 44

    # Separador
    r_sep = 4 + len(juan_rows) + 1
    ws.cell(row=r_sep, column=1, value="").border = None

    # --- MARÍA ---
    r_maria = r_sep + 1
    ws.cell(row=r_maria, column=1,
            value="PERFIL B — MARÍA RODRÍGUEZ (Capítulo II Sec. I · Actividad Profesional / Honorarios)")
    ws.cell(row=r_maria, column=1).font = FONT_H2

    maria_rows = [
        ("Ingresos cobrados por honorarios (anual)", 1_100_000, True),
        ("--- DEDUCCIONES DEL NEGOCIO (Art. 103 LISR) ---", None, False),
        ("Gastos deducibles indispensables del negocio", 250_000, True),
        ("Base gravable antes de deducciones personales", None, False),
        ("", None, False),
        ("--- DEDUCCIONES PERSONALES ---", None, False),
        ("Gastos médicos y hospitalarios (D01)", 45_000, True),
        ("Primas seguro gastos médicos (D07)", 20_000, True),
        ("Aportaciones PPR — bruto declarado (D06)", 85_000, True),
        ("PPR con límite Art. 151 Frac. V [MIN(bruto, 10% ing, 5SMG)]", None, False),
        ("Subtotal deducciones Arts. 151+185", None, False),
        ("Tope 15% sobre ingresos acumulables", None, False),
        ("Tope global (MIN entre 15% y 5 UMAs = $213,973.20)", None, False),
        ("Deducciones personales que aplican", None, False),
        ("", None, False),
        ("--- CÁLCULO ISR ---", None, False),
        ("Base gravable final (fila 4 − ded. personales)", None, False),
        ("LI aplicable (BUSCARV)", None, False),
        ("Excedente sobre LI", None, False),
        ("Tasa marginal", None, False),
        ("Impuesto marginal", None, False),
        ("Cuota fija (BUSCARV)", None, False),
        ("ISR determinado", None, False),
        ("Menos: pagos provisionales propios acreditables", 140_000, True),
        ("ISR A CARGO (+) / SALDO A FAVOR (−)", None, False),
    ]

    for j, (etiqueta, valor, es_input) in enumerate(maria_rows):
        r = r_maria + 1 + j
        ws.row_dimensions[r].height = 16
        c_label = ws.cell(row=r, column=1, value=etiqueta)
        c_label.font = FONT_BODY if etiqueta and not etiqueta.startswith("---") else FONT_H3
        c_label.alignment = ALIGN_LEFT
        c_label.border = BORDER_ALL if etiqueta and not etiqueta.startswith("---") and etiqueta != "" else None

        if es_input and valor is not None:
            input_cell(ws, r, 2, valor)
        elif etiqueta and not etiqueta.startswith("---") and etiqueta != "":
            blank_cell(ws, r, 2)

    ws.cell(row=r_maria + 2, column=3,
            value="Honorarios por servicios profesionales (Art. 100 LISR)").font = FONT_BODY_MUTED
    ws.cell(row=r_maria + 4, column=3,
            value="Renta de oficina, honorarios subcontratados, gastos con CFDI G03").font = FONT_BODY_MUTED
    ws.cell(row=r_maria + 11, column=3,
            value="10% × $1,100,000 = $110,000 > $85,000 → el límite es el bruto. 5SMG=$574,948 también > $85,000 → aplica bruto").font = FONT_BODY_MUTED
    ws.cell(row=r_maria + 25, column=3,
            value="Pagos provisionales acumulados enero–diciembre (Art. 106 LISR)").font = FONT_BODY_MUTED

    # ===== Solución =====
    ws = wb.create_sheet("Solución")
    set_column_widths(ws, [46, 22, 44])
    title_block(ws, "Solución — Declaración anual 2026 (valores calculados)", "")

    # --- JUAN SOLUCIÓN ---
    ws["A3"] = "PERFIL A — JUAN PÉREZ"
    ws["A3"].font = FONT_H2

    ing_juan = 650_000
    ret_juan = 122_000
    med_juan = 30_000
    seg_juan = 15_000
    ppr_bruto_juan = 25_000

    ppr_lim_juan = min(ppr_bruto_juan, ing_juan * 0.10, TOPE_SMG_5)
    subtotal_ded_juan = med_juan + seg_juan + ppr_lim_juan
    tope15_juan = ing_juan * 0.15
    tope_global_juan = min(tope15_juan, TOPE_UMA_5)
    ded_juan = min(subtotal_ded_juan, tope_global_juan)
    base_juan = ing_juan - ded_juan

    # Find bracket
    li_juan = cf_juan = tasa_juan = 0
    for li, ls, cf, tasa in TARIFA_ANUAL_2026:
        if li <= base_juan <= ls:
            li_juan, cf_juan, tasa_juan = li, cf, tasa
            break

    isr_juan = (base_juan - li_juan) * tasa_juan + cf_juan
    resultado_juan = isr_juan - ret_juan

    juan_sol = [
        ("Ingresos gravados por nómina (anual)", ing_juan, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Ingresos exentos (aguinaldo, prima vacacional)", 0, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Ingresos acumulables del ejercicio", ing_juan, '"$"#,##0.00', FILL_GREEN),
        ("", None, None, None),
        ("--- DEDUCCIONES PERSONALES ---", None, None, None),
        ("Gastos médicos y hospitalarios (D01)", med_juan, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Primas seguro gastos médicos (D07)", seg_juan, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Aportaciones PPR — bruto declarado (D06)", ppr_bruto_juan, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        (f"PPR con límite [MIN({ppr_bruto_juan:,.0f}, {ing_juan*0.10:,.0f}, {TOPE_SMG_5:,.0f})]", ppr_lim_juan, '"$"#,##0.00', None),
        ("Subtotal deducciones Arts. 151+185", subtotal_ded_juan, '"$"#,##0.00', None),
        (f"Tope 15% ({ing_juan:,.0f} × 15%)", tope15_juan, '"$"#,##0.00', None),
        ("Tope 5 UMAs ($213,973.20)", TOPE_UMA_5, '"$"#,##0.00', None),
        (f"Tope que aplica (MIN {tope15_juan:,.0f}, {TOPE_UMA_5:,.0f})", tope_global_juan, '"$"#,##0.00', None),
        ("Deducciones personales que aplican", ded_juan, '"$"#,##0.00', FILL_GREEN),
        ("", None, None, None),
        ("--- CÁLCULO ISR ---", None, None, None),
        ("Base gravable", base_juan, '"$"#,##0.00', None),
        ("LI aplicable", li_juan, '"$"#,##0.00', None),
        ("Excedente sobre LI", base_juan - li_juan, '"$"#,##0.00', None),
        ("Tasa marginal", tasa_juan, "0.00%", None),
        ("Impuesto marginal", (base_juan - li_juan) * tasa_juan, '"$"#,##0.00', None),
        ("Cuota fija", cf_juan, '"$"#,##0.00', None),
        ("ISR determinado", isr_juan, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Retenciones por patrón", ret_juan, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("ISR A CARGO (+) / SALDO A FAVOR (−)", resultado_juan, '"$"#,##0.00',
         FILL_GREEN if resultado_juan <= 0 else FILL_RED),
    ]

    for j, row_data in enumerate(juan_sol):
        etiqueta, valor, fmt, fill = row_data
        r = 4 + j
        c_label = ws.cell(row=r, column=1, value=etiqueta)
        c_label.alignment = ALIGN_LEFT
        if etiqueta and etiqueta.startswith("---"):
            c_label.font = FONT_H3
        else:
            c_label.font = FONT_BODY

        if valor is not None and fmt:
            c_val = ws.cell(row=r, column=2, value=valor)
            c_val.number_format = fmt
            c_val.alignment = ALIGN_RIGHT
            c_val.border = BORDER_ALL
            if fill:
                c_val.fill = fill
            if etiqueta and "ISR A CARGO" in etiqueta:
                c_val.font = FONT_RESULT
            else:
                c_val.font = FONT_BODY

    # Nota resultado Juan
    etiqueta_res = "→ Juan tiene SALDO A FAVOR" if resultado_juan <= 0 else "→ Juan paga en abril"
    ws.cell(row=4 + len(juan_sol), column=2, value=f"${abs(resultado_juan):,.2f} — {etiqueta_res}").font = FONT_NOTE

    # --- MARÍA SOLUCIÓN ---
    r_maria_sol = 4 + len(juan_sol) + 3
    ws.cell(row=r_maria_sol, column=1, value="PERFIL B — MARÍA RODRÍGUEZ").font = FONT_H2

    ing_maria = 1_100_000
    gastos_neg_maria = 250_000
    prov_maria = 140_000
    med_maria = 45_000
    seg_maria = 20_000
    ppr_bruto_maria = 85_000

    base_negocio_maria = ing_maria - gastos_neg_maria  # 850,000
    ppr_lim_maria = min(ppr_bruto_maria, ing_maria * 0.10, TOPE_SMG_5)  # min(85K, 110K, 574948) = 85K
    subtotal_ded_maria = med_maria + seg_maria + ppr_lim_maria  # 150,000
    tope15_maria = ing_maria * 0.15  # 165,000
    tope_global_maria = min(tope15_maria, TOPE_UMA_5)  # min(165K, 213973) = 165,000
    ded_maria = min(subtotal_ded_maria, tope_global_maria)  # min(150K, 165K) = 150,000
    base_final_maria = base_negocio_maria - ded_maria  # 850,000 - 150,000 = 700,000

    li_m = cf_m = tasa_m = 0
    for li, ls, cf, tasa in TARIFA_ANUAL_2026:
        if li <= base_final_maria <= ls:
            li_m, cf_m, tasa_m = li, cf, tasa
            break

    isr_maria = (base_final_maria - li_m) * tasa_m + cf_m
    resultado_maria = isr_maria - prov_maria

    maria_sol = [
        ("Ingresos cobrados por honorarios (anual)", ing_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("--- DEDUCCIONES DEL NEGOCIO ---", None, None, None),
        ("Gastos deducibles indispensables del negocio", gastos_neg_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Base antes de deducciones personales", base_negocio_maria, '"$"#,##0.00', FILL_GREEN),
        ("", None, None, None),
        ("--- DEDUCCIONES PERSONALES ---", None, None, None),
        ("Gastos médicos y hospitalarios (D01)", med_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Primas seguro gastos médicos (D07)", seg_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Aportaciones PPR — bruto declarado (D06)", ppr_bruto_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        (f"PPR con límite [MIN({ppr_bruto_maria:,.0f}, {ing_maria*0.10:,.0f}, {TOPE_SMG_5:,.0f})] → aplica bruto", ppr_lim_maria, '"$"#,##0.00', None),
        ("Subtotal deducciones Arts. 151+185", subtotal_ded_maria, '"$"#,##0.00', None),
        (f"Tope 15% ({ing_maria:,.0f} × 15%)", tope15_maria, '"$"#,##0.00', None),
        ("Tope 5 UMAs ($213,973.20)", TOPE_UMA_5, '"$"#,##0.00', None),
        (f"Tope que aplica (MIN {tope15_maria:,.0f}, {TOPE_UMA_5:,.0f})", tope_global_maria, '"$"#,##0.00', None),
        ("Deducciones personales que aplican", ded_maria, '"$"#,##0.00', FILL_GREEN),
        ("", None, None, None),
        ("--- CÁLCULO ISR ---", None, None, None),
        ("Base gravable final", base_final_maria, '"$"#,##0.00', None),
        ("LI aplicable", li_m, '"$"#,##0.00', None),
        ("Excedente sobre LI", base_final_maria - li_m, '"$"#,##0.00', None),
        ("Tasa marginal", tasa_m, "0.00%", None),
        ("Impuesto marginal", (base_final_maria - li_m) * tasa_m, '"$"#,##0.00', None),
        ("Cuota fija", cf_m, '"$"#,##0.00', None),
        ("ISR determinado", isr_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("Pagos provisionales acreditables", prov_maria, '"$"#,##0.00', FILL_PRIMARY_SOFT),
        ("ISR A CARGO (+) / SALDO A FAVOR (−)", resultado_maria, '"$"#,##0.00',
         FILL_GREEN if resultado_maria <= 0 else FILL_RED),
    ]

    for j, row_data in enumerate(maria_sol):
        etiqueta, valor, fmt, fill = row_data
        r = r_maria_sol + 1 + j
        c_label = ws.cell(row=r, column=1, value=etiqueta)
        c_label.alignment = ALIGN_LEFT
        if etiqueta and etiqueta.startswith("---"):
            c_label.font = FONT_H3
        else:
            c_label.font = FONT_BODY

        if valor is not None and fmt:
            c_val = ws.cell(row=r, column=2, value=valor)
            c_val.number_format = fmt
            c_val.alignment = ALIGN_RIGHT
            c_val.border = BORDER_ALL
            if fill:
                c_val.fill = fill
            if etiqueta and "ISR A CARGO" in etiqueta:
                c_val.font = FONT_RESULT
            else:
                c_val.font = FONT_BODY

    etiqueta_res_m = "→ María tiene SALDO A FAVOR" if resultado_maria <= 0 else "→ María paga en abril"
    ws.cell(row=r_maria_sol + 1 + len(maria_sol), column=2,
            value=f"${abs(resultado_maria):,.2f} — {etiqueta_res_m}").font = FONT_NOTE

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 44

    wb.save(path)
    print(f"✓ {path.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    here = Path(__file__).parent
    build_ejercicio_1(here / "01-deducciones-personales-tope.xlsx")
    build_ejercicio_2(here / "02-declaracion-anual-pf.xlsx")
    print("\nListo. 2 ejercicios generados.")


if __name__ == "__main__":
    main()
